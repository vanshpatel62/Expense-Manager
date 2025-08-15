from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum
from django.http import JsonResponse
from django.http import HttpResponse
import openpyxl
from .models import *
from .forms import *

def home(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    return render(request, 'expenses/login.html')

def register(request):
    if request.method == 'POST':
        form = UserRegisterForm(request.POST)
        if form.is_valid():
            try:
                user = form.save()
                UserProfile.objects.create(user=user, currency='INR')
                # Create default categories for new user (only if they don't exist globally)
                create_default_categories()
                username = form.cleaned_data.get('username')
                messages.success(request, f'Account created for {username}!')
                login(request, user)
                return redirect('dashboard')
            except Exception as e:
                messages.error(request, f'Error creating account: {str(e)}')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = UserRegisterForm()
    return render(request, 'expenses/register.html', {'form': form})

def create_default_categories():
    """Create default expense and income categories globally"""
    # Default expense categories
    expense_categories = [
        'food', 'transport', 'social_life', 'pets', 
        'apparel', 'health', 'education', 'other'
    ]
    
    # Default income categories
    income_categories = [
        'salary', 'allowance', 'bonus', 'other'
    ]
    
    # Create expense categories
    for category_name in expense_categories:
        ExpenseCategory.objects.get_or_create(
            name=category_name
        )
    
    # Create income categories
    for category_name in income_categories:
        IncomeCategory.objects.get_or_create(
            name=category_name
        )

def user_login(request):
    if request.method == 'POST':
        form = LoginForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('dashboard')
            else:
                messages.error(request, "Invalid username or password")
        else:
            messages.error(request, "Invalid username or password")
    else:
        form = LoginForm()
    return render(request, 'expenses/login.html', {'form': form})

@login_required
def user_logout(request):
    logout(request)
    return redirect('home')

@login_required
def dashboard(request):
    user = request.user
    accounts = Account.objects.filter(user=user)
    transactions = Transaction.objects.filter(account__user=user).order_by('-transaction_date')[:5]

    total_balance = sum(account.current_balance for account in accounts)

    # Sum income from Transactions
    total_income = Transaction.objects.filter(
        account__user=user, transaction_type='income'
    ).aggregate(total=Sum('amount'))['total'] or 0

    # Sum expense from Transactions
    total_expense_trans = Transaction.objects.filter(
        account__user=user, transaction_type='expense'
    ).aggregate(total=Sum('amount'))['total'] or 0

    # Sum expense from Expense model
    total_expense_exp = Expense.objects.filter(user=user).aggregate(total=Sum('amount'))['total'] or 0

    # Combine expenses from both models
    total_expense = total_expense_trans + total_expense_exp

    # Get expense chart data for dashboard
    expense_chart_data = get_expense_chart_data(user)

    context = {
        'accounts': accounts,
        'transactions': transactions,
        'total_balance': total_balance,
        'total_income': total_income,
        'total_expense': total_expense,
        'expense_chart_data': expense_chart_data,
    }
    return render(request, 'expenses/dashboard.html', context)

def get_expense_chart_data(user):
    """Helper function to get expense chart data"""
    # Get expense data from Transaction model
    expense_transactions = Transaction.objects.filter(
        account__user=user,
        transaction_type='expense',
        expense_category__isnull=False
    ).values('expense_category__name').annotate(
        total=Sum('amount')
    ).order_by('-total')
    
    # Get expense data from Expense model
    expense_records = Expense.objects.filter(
        user=user
    ).values('category__name').annotate(
        total=Sum('amount')
    ).order_by('-total')
    
    # Combine data from both models
    chart_data = {}
    
    # Add transaction expenses
    for item in expense_transactions:
        category_name = item['expense_category__name']
        if category_name in chart_data:
            chart_data[category_name] += float(item['total'])
        else:
            chart_data[category_name] = float(item['total'])
    
    # Add expense model data
    for item in expense_records:
        category_name = item['category__name']
        if category_name in chart_data:
            chart_data[category_name] += float(item['total'])
        else:
            chart_data[category_name] = float(item['total'])
    
    # Prepare data for Chart.js
    labels = list(chart_data.keys())
    data = list(chart_data.values())
    
    # Generate colors for each category
    colors = [
        '#FF6384', '#36A2EB', '#FFCE56', '#4BC0C0', 
        '#9966FF', '#FF9F40', '#FF6384', '#C9CBCF'
    ]
    
    chart_colors = colors[:len(labels)]
    
    return {
        'labels': labels,
        'data': data,
        'colors': chart_colors,
    }

@login_required
def accounts_view(request):
    accounts = Account.objects.filter(user=request.user)
    return render(request, 'expenses/accounts.html', {'accounts': accounts})

@login_required
def add_account(request):
    if request.method == 'POST':
        form = AccountForm(request.POST)
        if form.is_valid():
            account = form.save(commit=False)
            account.user = request.user
            account.save()
            messages.success(request, 'Account added successfully!')
            return redirect('accounts')
    else:
        form = AccountForm()
    return render(request, 'expenses/add_account.html', {'form': form})

@login_required
def account_detail(request, pk):
    account = get_object_or_404(Account, pk=pk, user=request.user)
    transactions = Transaction.objects.filter(account=account).order_by('-transaction_date')
    return render(request, 'expenses/account_detail.html', {
        'account': account,
        'transactions': transactions
    })

@login_required
def transactions_view(request):
    transactions = Transaction.objects.filter(account__user=request.user).order_by('-transaction_date')
    return render(request, 'expenses/transactions.html', {'transactions': transactions})

@login_required
def add_transaction(request):
    if request.method == 'POST':
        form = TransactionForm(request.POST, user=request.user)
        if form.is_valid():
            transaction = form.save(commit=False)
            transaction.account = form.cleaned_data['account']
            
            # Check if user is trying to make a transfer but only has one account
            if transaction.transaction_type == 'transfer':
                user_accounts = Account.objects.filter(user=request.user)
                if user_accounts.count() == 1:
                    messages.error(request, 'You need at least two accounts to make a transfer. Please add another account first.')
                    return render(request, 'expenses/add_transaction.html', {'form': form})
            
            # Handle different transaction types
            if transaction.transaction_type == 'income':
                transaction.income_category = form.cleaned_data['income_category']
                # Update account balance
                transaction.account.current_balance += transaction.amount
                transaction.account.save()
                
            elif transaction.transaction_type == 'expense':
                transaction.expense_category = form.cleaned_data['expense_category']
                # Update account balance
                transaction.account.current_balance -= transaction.amount
                transaction.account.save()
                
            elif transaction.transaction_type == 'transfer':
                transaction.to_account = form.cleaned_data['to_account']
                
                # Check if transferring to the same account
                if transaction.account == transaction.to_account:
                    messages.error(request, 'Cannot transfer money to the same account. Please select a different destination account.')
                    return render(request, 'expenses/add_transaction.html', {'form': form})
                
                # Update both account balances
                transaction.account.current_balance -= transaction.amount
                transaction.to_account.current_balance += transaction.amount
                transaction.account.save()
                transaction.to_account.save()
            
            transaction.save()
            messages.success(request, 'Transaction added successfully!')
            return redirect('transactions')
    else:
        form = TransactionForm(user=request.user)
    return render(request, 'expenses/add_transaction.html', {'form': form})

@login_required
def edit_transaction(request, pk):
    transaction = get_object_or_404(Transaction, pk=pk, account__user=request.user)
    if request.method == 'POST':
        form = TransactionForm(request.POST, instance=transaction, user=request.user)
        if form.is_valid():
            # Check if user is trying to make a transfer but only has one account
            if form.cleaned_data['transaction_type'] == 'transfer':
                user_accounts = Account.objects.filter(user=request.user)
                if user_accounts.count() == 1:
                    messages.error(request, 'You need at least two accounts to make a transfer. Please add another account first.')
                    return render(request, 'expenses/edit_transaction.html', {'form': form})
            
            # Handle balance adjustments for existing transaction
            old_amount = transaction.amount
            old_type = transaction.transaction_type
            
            transaction = form.save(commit=False)
            transaction.account = form.cleaned_data['account']
            
            # Revert old transaction effects
            if old_type == 'income':
                transaction.account.current_balance -= old_amount
            elif old_type == 'expense':
                transaction.account.current_balance += old_amount
            elif old_type == 'transfer' and transaction.to_account:
                transaction.account.current_balance += old_amount
                transaction.to_account.current_balance -= old_amount
                transaction.to_account.save()
            
            # Apply new transaction effects
            if transaction.transaction_type == 'income':
                transaction.income_category = form.cleaned_data['income_category']
                transaction.account.current_balance += transaction.amount
            elif transaction.transaction_type == 'expense':
                transaction.expense_category = form.cleaned_data['expense_category']
                transaction.account.current_balance -= transaction.amount
            elif transaction.transaction_type == 'transfer':
                transaction.to_account = form.cleaned_data['to_account']
                
                # Check if transferring to the same account
                if transaction.account == transaction.to_account:
                    messages.error(request, 'Cannot transfer money to the same account. Please select a different destination account.')
                    return render(request, 'expenses/edit_transaction.html', {'form': form})
                
                transaction.account.current_balance -= transaction.amount
                transaction.to_account.current_balance += transaction.amount
                transaction.to_account.save()
            
            transaction.account.save()
            transaction.save()
            messages.success(request, 'Transaction updated successfully!')
            return redirect('transactions')
    else:
        form = TransactionForm(instance=transaction, user=request.user)
    return render(request, 'expenses/edit_transaction.html', {'form': form})

@login_required
def delete_transaction(request, pk):
    transaction = get_object_or_404(Transaction, pk=pk, account__user=request.user)
    if request.method == 'POST':
        # Revert balance changes
        if transaction.transaction_type == 'income':
            transaction.account.current_balance -= transaction.amount
        elif transaction.transaction_type == 'expense':
            transaction.account.current_balance += transaction.amount
        elif transaction.transaction_type == 'transfer' and transaction.to_account:
            transaction.account.current_balance += transaction.amount
            transaction.to_account.current_balance -= transaction.amount
            transaction.to_account.save()
        
        transaction.account.save()
        transaction.delete()
        messages.success(request, 'Transaction deleted successfully!')
        return redirect('transactions')
    return render(request, 'expenses/delete_transaction.html', {'transaction': transaction})

@login_required
def categories_view(request):
    expense_categories = ExpenseCategory.objects.all()
    income_categories = IncomeCategory.objects.all()
    return render(request, 'expenses/categories.html', {
        'expense_categories': expense_categories,
        'income_categories': income_categories
    })

@login_required
def add_expense_category(request):
    if request.method == 'POST':
        form = ExpenseCategoryForm(request.POST)
        if form.is_valid():
            category = form.save()
            messages.success(request, 'Expense category added successfully!')
            return redirect('categories')
    else:
        form = ExpenseCategoryForm()
    return render(request, 'expenses/add_expense_category.html', {'form': form})

@login_required
def add_income_category(request):
    if request.method == 'POST':
        form = IncomeCategoryForm(request.POST)
        if form.is_valid():
            category = form.save()
            messages.success(request, 'Income category added successfully!')
            return redirect('categories')
    else:
        form = IncomeCategoryForm()
    return render(request, 'expenses/add_income_category.html', {'form': form})

@login_required
def expenses_view(request):
    expenses = Expense.objects.filter(user=request.user).order_by('-date')
    return render(request, 'expenses/expenses.html', {'expenses': expenses})

@login_required
def add_expense(request):
    if request.method == 'POST':
        form = ExpenseForm(request.POST, user=request.user)
        if form.is_valid():
            expense = form.save(commit=False)
            expense.user = request.user
            expense.save()
            messages.success(request, 'Expense added successfully!')
            return redirect('expenses')
    else:
        form = ExpenseForm(user=request.user)
    return render(request, 'expenses/add_expense.html', {'form': form})

@login_required
def edit_expense(request, pk):
    expense = get_object_or_404(Expense, pk=pk, user=request.user)
    if request.method == 'POST':
        form = ExpenseForm(request.POST, instance=expense, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Expense updated successfully!')
            return redirect('expenses')
    else:
        form = ExpenseForm(instance=expense, user=request.user)
    return render(request, 'expenses/edit_expense.html', {'form': form})

@login_required
def delete_expense(request, pk):
    expense = get_object_or_404(Expense, pk=pk, user=request.user)
    if request.method == 'POST':
        expense.delete()
        messages.success(request, 'Expense deleted successfully!')
        return redirect('expenses')
    return render(request, 'expenses/delete_expense.html', {'expense': expense})

@login_required
def profile_view(request):
    profile, created = UserProfile.objects.get_or_create(
        user=request.user,
        defaults={'currency': 'INR'}
    )
    
    # If profile was created with default currency, ensure it's INR
    if created and profile.currency != 'INR':
        profile.currency = 'INR'
        profile.save()
    
    if request.method == 'POST':
        form = UserProfileForm(request.POST, instance=profile)
        if form.is_valid():
            form.save()
            messages.success(request, 'Profile updated successfully!')
            return redirect('profile')
    else:
        form = UserProfileForm(instance=profile)
    return render(request, 'expenses/profile.html', {'form': form})

@login_required
def expense_charts_view(request):
    """View for displaying expense category pie charts"""
    user = request.user
    
    # Get expense chart data using helper function
    chart_data = get_expense_chart_data(user)
    
    context = {
        'labels': chart_data['labels'],
        'data': chart_data['data'],
        'colors': chart_data['colors'],
        'total_expense': sum(chart_data['data'])
    }
    
    return render(request, 'expenses/expense_charts.html', context)

@login_required
def     export_transactions_excel(request):
    from .models import Transaction
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Transactions'
    headers = ['Date', 'Type', 'Account', 'Amount', 'Description', 'Expense Category', 'Income Category', 'To Account']
    ws.append(headers)

    # Make header bold
    from openpyxl.styles import Font
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = Font(bold=True)

    transactions = Transaction.objects.filter(account__user=request.user).order_by('-transaction_date')
    for t in transactions:
        ws.append([
            t.transaction_date.strftime('%Y-%m-%d') if t.transaction_date else '',
            t.get_transaction_type_display(),
            str(t.account),
            float(t.amount),
            t.description,
            str(t.expense_category) if t.expense_category else '',
            str(t.income_category) if t.income_category else '',
            str(t.to_account) if t.to_account else '',
        ])

    # Auto-fit column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    # Format amount as currency
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=4, max_row=ws.max_row):
        for cell in row:
            cell.number_format = '#,##0.00'

    # Calculate totals
    total_expense = sum(float(t.amount) for t in transactions if t.transaction_type == 'expense')
    total_income = sum(float(t.amount) for t in transactions if t.transaction_type == 'income')

    # Leave a blank row, then add totals
    ws.append([])
    ws.append(['', 'Total Expense', '', total_expense, '', '', '', ''])
    ws.append(['', 'Total Income', '', total_income, '', '', '', ''])

    # Format the total rows' amount as currency
    for row in ws.iter_rows(min_row=ws.max_row-1, max_row=ws.max_row, min_col=4, max_col=4):
        for cell in row:
            cell.number_format = '#,##0.00'

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Add autofilter
    ws.auto_filter.ref = ws.dimensions

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"{request.user.username}-Transactions.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response
